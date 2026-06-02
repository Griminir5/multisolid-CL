from __future__ import annotations

import json
import math
from pathlib import Path
import shutil
from dataclasses import dataclass
from typing import Any, Mapping

from daetools.pyDAE import Constant, Exp, Max, Min

from pyUnits import J, K, Pa, m, mol, s

from ..properties import PROPERTY_REGISTRY
from . import KineticsContext, register_kinetics_hook


GAS_CONSTANT_J_PER_MOL_K = 8.31446261815324
PRESSURE_PA_PER_KPA = 1.0e3
# The Ortiz shrinking-core form becomes numerically singular as X -> 1.
# For packed-bed comparison runs we do not resolve the last 0.01% of residual
# oxygen accessibility, so keep a small physically negligible floor that
# prevents the solver from spending most of its time on the asymptotic tail.
MIN_ONE_MINUS_CONVERSION = 1.0e-4
MIN_GAS_CONCENTRATION_MOL_PER_M3 = 1.0e-12
MIN_PARTIAL_PRESSURE_KPA = 1.0e-6
WGS_REACTANT_GATE_KPA = 1.0e-6
MIN_SOLID_CAPACITY_MOL_PER_M3 = 1.0e-18
ILMENITE_ACTIVE_GATE_FRACTION = 1.0e-3
STANDARD_TEMPERATURE_K = 298.15


# Ortiz et al., Energy Technology 2016, Table 3.
# The fitted redox law is an integrated-rate-of-reduction/oxidation (IRoR)
# shrinking-core model for the whole ilmenite particle, not separate fits for
# each resolved phase. This module keeps the paper rate expression intact and
# only uses phase shares to map the total particle rate onto the explicit
# solver reactions.
ILMENITE_RATE_COEFFICIENTS_M_PER_S = {
    "h2_reduction": 8.944,
    "co_reduction": 0.193,
    "o2_oxidation": 0.438,
}
ILMENITE_DIFFUSIVITY_COEFFICIENTS_M2_PER_S = {
    "h2_reduction": 42.0,
    # Table 3 prints D0 = 5.6 x 10^5 cm^2/s for CO reduction. Using that
    # literal value does not reproduce the paper's own Figure 3 model lines:
    # the full Eq. (5) solve underpredicts all six CO traces substantially.
    # A factor-of-10 larger D0 is the minimal correction that restores
    # consistency between Table 3 and Figure 3 while leaving the H2/O2
    # branches unchanged, so this is treated as a figure-consistent paper
    # correction rather than a new fit.
    "co_reduction": 560.0,
    "o2_oxidation": 2.84e-2,
}
ILMENITE_ACTIVATION_ENERGIES_J_PER_MOL = {
    "h2_reduction": 45.5e3,
    "co_reduction": 30.6e3,
    "o2_oxidation": 5.3e3,
}
ILMENITE_DIFFUSION_ACTIVATION_ENERGIES_J_PER_MOL = {
    "h2_reduction": 99.7e3,
    "co_reduction": 154.5e3,
    "o2_oxidation": 4.4e3,
}
ILMENITE_SOLID_CONCENTRATIONS_MOL_PER_M3 = {
    "h2_reduction": 7.2e3,
    "co_reduction": 7.2e3,
    # Table 3 doubles Cs for oxidation, so the O2 pathway is kept on the same
    # one-oxygen-site basis without an extra stoichiometric b factor.
    "o2_oxidation": 14.4e3,
}
ILMENITE_PARTICLE_RADIUS_M = 2.13e-3
ILMENITE_CONVERSION_DECAY = {
    "h2_reduction": 3.5,
    "co_reduction": 5.0,
    "o2_oxidation": 0.0,
}
ILMENITE_REACTION_ORDER = {
    "h2_reduction": 1.0,
    "co_reduction": 1.0,
    "o2_oxidation": 1.0,
}
ILMENITE_STOICHIOMETRIC_FACTOR = {
    "h2_reduction": 1.0,
    "co_reduction": 1.0,
    "o2_oxidation": 1.0,
}


ILMENITE_WGS_PREEXPONENTIAL_MOL_PER_G_S = 2.52e-5
ILMENITE_WGS_ACTIVATION_ENERGY_J_PER_MOL = 22.0e3
ILMENITE_WGS_REACTION_ORDERS = {
    "CO": 0.65,
    "H2O": 0.3,
    "CO2": -0.2,
    "H2": -0.7,
}


FE2O3_MW_KG_PER_MOL = PROPERTY_REGISTRY.get_record("Fe2O3").mw
FEO_MW_KG_PER_MOL = PROPERTY_REGISTRY.get_record("FeO").mw
FE2TIO4_MW_KG_PER_MOL = PROPERTY_REGISTRY.get_record("Fe2TiO4").mw
FE2TIO5_MW_KG_PER_MOL = PROPERTY_REGISTRY.get_record("Fe2TiO5").mw
FE3O4_MW_KG_PER_MOL = PROPERTY_REGISTRY.get_record("Fe3O4").mw
TIO2_MW_KG_PER_MOL = PROPERTY_REGISTRY.get_record("TiO2").mw
FETIO3_MW_KG_PER_MOL = FEO_MW_KG_PER_MOL + TIO2_MW_KG_PER_MOL


if any(
    value is None
    for value in (
        FE2O3_MW_KG_PER_MOL,
        FEO_MW_KG_PER_MOL,
        FE2TIO4_MW_KG_PER_MOL,
        FE2TIO5_MW_KG_PER_MOL,
        FE3O4_MW_KG_PER_MOL,
        TIO2_MW_KG_PER_MOL,
    )
):
    raise ValueError("Ilmenite kinetics require Fe/Ti species molecular weights in PROPERTY_REGISTRY.")


ILMENITE_REACTION_STOICHIOMETRIES: Mapping[str, Mapping[str, float]] = {
    "ilmenite_fe2tio5_h2_reduction_ortiz_2016": {
        "Fe2TiO5": -1.0,
        "H2": -1.0,
        "Fe2TiO4": 1.0,
        "H2O": 1.0,
    },
    "ilmenite_fe2o3_h2_reduction_ortiz_2016": {
        "Fe2O3": -1.0,
        "H2": -1.0,
        "FeO": 2.0,
        "H2O": 1.0,
    },
    "ilmenite_fe2tio5_co_reduction_ortiz_2016": {
        "Fe2TiO5": -1.0,
        "CO": -1.0,
        "Fe2TiO4": 1.0,
        "CO2": 1.0,
    },
    "ilmenite_fe2o3_co_reduction_ortiz_2016": {
        "Fe2O3": -1.0,
        "CO": -1.0,
        "FeO": 2.0,
        "CO2": 1.0,
    },
    "ilmenite_fe2tio4_o2_oxidation_ortiz_2016": {
        "Fe2TiO4": -1.0,
        "O2": -0.5,
        "Fe2TiO5": 1.0,
    },
    "ilmenite_feo_o2_oxidation_ortiz_2016": {
        "FeO": -2.0,
        "O2": -0.5,
        "Fe2O3": 1.0,
    },
    "ilmenite_wgs_ortiz_2016": {
        "CO": -1.0,
        "H2O": -1.0,
        "CO2": 1.0,
        "H2": 1.0,
    },
}


@dataclass(frozen=True)
class IlmeniteRedoxTerms:
    temperature_k: Any
    pressure_kpa: Any
    c_h2_mol_per_m3: Any
    c_co_mol_per_m3: Any
    c_o2_mol_per_m3: Any
    p_h2_kpa: Any
    p_h2o_kpa: Any
    p_co_kpa: Any
    p_co2_kpa: Any
    c_fe2tio5_mol_per_m3: Any
    c_fe2tio4_mol_per_m3: Any
    c_fe2o3_mol_per_m3: Any
    c_feo_mol_per_m3: Any
    oxygen_capacity_mol_per_m3: Any
    oxidized_inventory_mol_per_m3: Any
    oxidation_demand_mol_per_m3: Any
    x_reduction: Any
    x_oxidation: Any
    catalyst_mass_density_kg_per_m3: Any


def partial_pressure_value(total_pressure_pa: float, mole_fraction: float) -> float:
    return total_pressure_pa * mole_fraction


def partial_pressure_kpa_value(total_pressure_pa: float, mole_fraction: float) -> float:
    return partial_pressure_value(total_pressure_pa, mole_fraction) / PRESSURE_PA_PER_KPA


def gas_phase_concentration_value(*, total_pressure_pa: float, mole_fraction: float, temperature_k: float) -> float:
    return max(total_pressure_pa, 0.0) * max(mole_fraction, 0.0) / (GAS_CONSTANT_J_PER_MOL_K * temperature_k)


def catalyst_mass_density_value(
    *,
    c_fe2tio5_mol_per_m3: float = 0.0,
    c_fe2tio4_mol_per_m3: float = 0.0,
    c_fe2o3_mol_per_m3: float = 0.0,
    c_feo_mol_per_m3: float = 0.0,
    c_fe3o4_mol_per_m3: float = 0.0,
    c_tio2_mol_per_m3: float = 0.0,
) -> float:
    return (
        max(c_fe2tio5_mol_per_m3, 0.0) * FE2TIO5_MW_KG_PER_MOL
        + max(c_fe2tio4_mol_per_m3, 0.0) * FE2TIO4_MW_KG_PER_MOL
        + max(c_fe2o3_mol_per_m3, 0.0) * FE2O3_MW_KG_PER_MOL
        + max(c_feo_mol_per_m3, 0.0) * FEO_MW_KG_PER_MOL
        + max(c_fe3o4_mol_per_m3, 0.0) * FE3O4_MW_KG_PER_MOL
        + max(c_tio2_mol_per_m3, 0.0) * TIO2_MW_KG_PER_MOL
    )


def oxygen_capacity_value(
    *,
    c_fe2tio5_mol_per_m3: float,
    c_fe2tio4_mol_per_m3: float,
    c_fe2o3_mol_per_m3: float,
    c_feo_mol_per_m3: float,
) -> float:
    return max(c_fe2tio5_mol_per_m3, 0.0) + max(c_fe2tio4_mol_per_m3, 0.0) + max(c_fe2o3_mol_per_m3, 0.0) + 0.5 * max(c_feo_mol_per_m3, 0.0)


def oxidized_inventory_value(
    *,
    c_fe2tio5_mol_per_m3: float,
    c_fe2o3_mol_per_m3: float,
) -> float:
    return max(c_fe2tio5_mol_per_m3, 0.0) + max(c_fe2o3_mol_per_m3, 0.0)


def oxidation_demand_value(
    *,
    c_fe2tio4_mol_per_m3: float,
    c_feo_mol_per_m3: float,
) -> float:
    return max(c_fe2tio4_mol_per_m3, 0.0) + 0.5 * max(c_feo_mol_per_m3, 0.0)


def reduction_conversion_value(
    *,
    c_fe2tio5_mol_per_m3: float,
    c_fe2tio4_mol_per_m3: float,
    c_fe2o3_mol_per_m3: float,
    c_feo_mol_per_m3: float,
) -> float:
    capacity = max(
        oxygen_capacity_value(
            c_fe2tio5_mol_per_m3=c_fe2tio5_mol_per_m3,
            c_fe2tio4_mol_per_m3=c_fe2tio4_mol_per_m3,
            c_fe2o3_mol_per_m3=c_fe2o3_mol_per_m3,
            c_feo_mol_per_m3=c_feo_mol_per_m3,
        ),
        MIN_SOLID_CAPACITY_MOL_PER_M3,
    )
    oxidized = oxidized_inventory_value(
        c_fe2tio5_mol_per_m3=c_fe2tio5_mol_per_m3,
        c_fe2o3_mol_per_m3=c_fe2o3_mol_per_m3,
    )
    return 1.0 - oxidized / capacity


def oxidation_conversion_value(
    *,
    c_fe2tio5_mol_per_m3: float,
    c_fe2tio4_mol_per_m3: float,
    c_fe2o3_mol_per_m3: float,
    c_feo_mol_per_m3: float,
) -> float:
    return 1.0 - reduction_conversion_value(
        c_fe2tio5_mol_per_m3=c_fe2tio5_mol_per_m3,
        c_fe2tio4_mol_per_m3=c_fe2tio4_mol_per_m3,
        c_fe2o3_mol_per_m3=c_fe2o3_mol_per_m3,
        c_feo_mol_per_m3=c_feo_mol_per_m3,
    )


def reaction_constant_value(rate_key: str, *, temperature_k: float) -> float:
    return ILMENITE_RATE_COEFFICIENTS_M_PER_S[rate_key] * math.exp(
        -ILMENITE_ACTIVATION_ENERGIES_J_PER_MOL[rate_key] / (GAS_CONSTANT_J_PER_MOL_K * temperature_k)
    )


def diffusivity_value(rate_key: str, *, temperature_k: float, conversion: float) -> float:
    return ILMENITE_DIFFUSIVITY_COEFFICIENTS_M2_PER_S[rate_key] * math.exp(
        -ILMENITE_DIFFUSION_ACTIVATION_ENERGIES_J_PER_MOL[rate_key] / (GAS_CONSTANT_J_PER_MOL_K * temperature_k)
    ) * math.exp(-ILMENITE_CONVERSION_DECAY[rate_key] * conversion)


def mixed_control_conversion_rate_value(
    rate_key: str,
    *,
    temperature_k: float,
    gas_concentration_mol_per_m3: float,
    conversion: float,
) -> float:
    conversion_safe = min(max(conversion, 0.0), 1.0 - MIN_ONE_MINUS_CONVERSION)
    one_minus_x = 1.0 - conversion_safe
    denominator = (
        (1.0 / reaction_constant_value(rate_key, temperature_k=temperature_k)) * one_minus_x ** (-2.0 / 3.0)
        + (ILMENITE_PARTICLE_RADIUS_M / diffusivity_value(rate_key, temperature_k=temperature_k, conversion=conversion_safe))
        * (one_minus_x ** (-1.0 / 3.0) - 1.0)
    )
    return (
        3.0
        * ILMENITE_STOICHIOMETRIC_FACTOR[rate_key]
        * max(gas_concentration_mol_per_m3, 0.0) ** ILMENITE_REACTION_ORDER[rate_key]
        / (ILMENITE_PARTICLE_RADIUS_M * ILMENITE_SOLID_CONCENTRATIONS_MOL_PER_M3[rate_key] * denominator)
    )


def h2_reduction_site_rate_value(
    *,
    temperature_k: float,
    c_h2_mol_per_m3: float,
    c_fe2tio5_mol_per_m3: float,
    c_fe2tio4_mol_per_m3: float,
    c_fe2o3_mol_per_m3: float,
    c_feo_mol_per_m3: float,
) -> float:
    capacity = oxygen_capacity_value(
        c_fe2tio5_mol_per_m3=c_fe2tio5_mol_per_m3,
        c_fe2tio4_mol_per_m3=c_fe2tio4_mol_per_m3,
        c_fe2o3_mol_per_m3=c_fe2o3_mol_per_m3,
        c_feo_mol_per_m3=c_feo_mol_per_m3,
    )
    x_reduction = reduction_conversion_value(
        c_fe2tio5_mol_per_m3=c_fe2tio5_mol_per_m3,
        c_fe2tio4_mol_per_m3=c_fe2tio4_mol_per_m3,
        c_fe2o3_mol_per_m3=c_fe2o3_mol_per_m3,
        c_feo_mol_per_m3=c_feo_mol_per_m3,
    )
    return capacity * mixed_control_conversion_rate_value(
        "h2_reduction",
        temperature_k=temperature_k,
        gas_concentration_mol_per_m3=c_h2_mol_per_m3,
        conversion=x_reduction,
    )


def co_reduction_site_rate_value(
    *,
    temperature_k: float,
    c_co_mol_per_m3: float,
    c_fe2tio5_mol_per_m3: float,
    c_fe2tio4_mol_per_m3: float,
    c_fe2o3_mol_per_m3: float,
    c_feo_mol_per_m3: float,
) -> float:
    capacity = oxygen_capacity_value(
        c_fe2tio5_mol_per_m3=c_fe2tio5_mol_per_m3,
        c_fe2tio4_mol_per_m3=c_fe2tio4_mol_per_m3,
        c_fe2o3_mol_per_m3=c_fe2o3_mol_per_m3,
        c_feo_mol_per_m3=c_feo_mol_per_m3,
    )
    x_reduction = reduction_conversion_value(
        c_fe2tio5_mol_per_m3=c_fe2tio5_mol_per_m3,
        c_fe2tio4_mol_per_m3=c_fe2tio4_mol_per_m3,
        c_fe2o3_mol_per_m3=c_fe2o3_mol_per_m3,
        c_feo_mol_per_m3=c_feo_mol_per_m3,
    )
    return capacity * mixed_control_conversion_rate_value(
        "co_reduction",
        temperature_k=temperature_k,
        gas_concentration_mol_per_m3=c_co_mol_per_m3,
        conversion=x_reduction,
    )


def o2_oxidation_site_rate_value(
    *,
    temperature_k: float,
    c_o2_mol_per_m3: float,
    c_fe2tio5_mol_per_m3: float,
    c_fe2tio4_mol_per_m3: float,
    c_fe2o3_mol_per_m3: float,
    c_feo_mol_per_m3: float,
) -> float:
    capacity = oxygen_capacity_value(
        c_fe2tio5_mol_per_m3=c_fe2tio5_mol_per_m3,
        c_fe2tio4_mol_per_m3=c_fe2tio4_mol_per_m3,
        c_fe2o3_mol_per_m3=c_fe2o3_mol_per_m3,
        c_feo_mol_per_m3=c_feo_mol_per_m3,
    )
    x_oxidation = oxidation_conversion_value(
        c_fe2tio5_mol_per_m3=c_fe2tio5_mol_per_m3,
        c_fe2tio4_mol_per_m3=c_fe2tio4_mol_per_m3,
        c_fe2o3_mol_per_m3=c_fe2o3_mol_per_m3,
        c_feo_mol_per_m3=c_feo_mol_per_m3,
    )
    return capacity * mixed_control_conversion_rate_value(
        "o2_oxidation",
        temperature_k=temperature_k,
        gas_concentration_mol_per_m3=c_o2_mol_per_m3,
        conversion=x_oxidation,
    )


def equilibrium_constant_wgs_value(temperature_k: float) -> float:
    z_value = 1000.0 / temperature_k - 1.0
    return math.exp(z_value * (z_value * (0.63508 - 0.29353 * z_value) + 4.1778) + 0.31688)


def wgs_rate_constant_value(*, temperature_k: float) -> float:
    return ILMENITE_WGS_PREEXPONENTIAL_MOL_PER_G_S * math.exp(
        -ILMENITE_WGS_ACTIVATION_ENERGY_J_PER_MOL / (GAS_CONSTANT_J_PER_MOL_K * temperature_k)
    )


def wgs_rate_value(
    *,
    temperature_k: float,
    p_co_kpa: float,
    p_h2o_kpa: float,
    p_co2_kpa: float,
    p_h2_kpa: float,
    catalyst_mass_density_kg_per_m3: float,
) -> float:
    p_co_available = max(p_co_kpa, 0.0)
    p_h2o_available = max(p_h2o_kpa, 0.0)
    p_co2_safe = max(p_co2_kpa, MIN_PARTIAL_PRESSURE_KPA)
    p_h2_safe = max(p_h2_kpa, MIN_PARTIAL_PRESSURE_KPA)
    p_co_safe = max(p_co_available, MIN_PARTIAL_PRESSURE_KPA)
    p_h2o_safe = max(p_h2o_available, MIN_PARTIAL_PRESSURE_KPA)
    reactant_gate = (
        p_co_available / (p_co_available + WGS_REACTANT_GATE_KPA)
        * p_h2o_available / (p_h2o_available + WGS_REACTANT_GATE_KPA)
    )
    equilibrium = equilibrium_constant_wgs_value(temperature_k)
    beta = (p_co2_safe * p_h2_safe) / (equilibrium * p_co_safe * p_h2o_safe)
    return (
        reactant_gate
        * wgs_rate_constant_value(temperature_k=temperature_k)
        * p_co_safe ** ILMENITE_WGS_REACTION_ORDERS["CO"]
        * p_h2o_safe ** ILMENITE_WGS_REACTION_ORDERS["H2O"]
        * p_co2_safe ** ILMENITE_WGS_REACTION_ORDERS["CO2"]
        * p_h2_safe ** ILMENITE_WGS_REACTION_ORDERS["H2"]
        * (1.0 - beta)
        * catalyst_mass_density_kg_per_m3
        * 1.0e3
    )


def reaction_enthalpy_value(reaction_id: str, *, temperature_k: float = STANDARD_TEMPERATURE_K) -> float:
    stoichiometry = ILMENITE_REACTION_STOICHIOMETRIES[reaction_id]
    return float(
        sum(
            coefficient * PROPERTY_REGISTRY.enthalpy_value(species_id, temperature_k)
            for species_id, coefficient in stoichiometry.items()
        )
    )


def reaction_enthalpy_expression(reaction_id: str, temperature):
    stoichiometry = ILMENITE_REACTION_STOICHIOMETRIES[reaction_id]
    expression = Constant(0.0 * J / mol)
    for species_id, coefficient in stoichiometry.items():
        expression = expression + Constant(coefficient) * PROPERTY_REGISTRY.enthalpy_expression(species_id, temperature)
    return expression


ILMENITE_REACTION_STANDARD_HEATS_J_PER_MOL = {
    reaction_id: reaction_enthalpy_value(reaction_id)
    for reaction_id in ILMENITE_REACTION_STOICHIOMETRIES
}


def _temperature_k_expression(temperature) -> Any:
    return temperature / Constant(1.0 * K)


def _pressure_pa_expression(pressure) -> Any:
    return Max(pressure / Constant(1.0 * Pa), Constant(0.0))


def _pressure_kpa_expression(pressure) -> Any:
    return _pressure_pa_expression(pressure) / Constant(PRESSURE_PA_PER_KPA)


def _positive_part_expression(expression):
    return Max(expression, Constant(0.0))


def _bounded_conversion_expression(conversion):
    return Min(Constant(1.0) - Constant(MIN_ONE_MINUS_CONVERSION), Max(conversion, Constant(0.0)))


def _safe_division(numerator, denominator):
    return numerator / Max(denominator, Constant(MIN_SOLID_CAPACITY_MOL_PER_M3))


def _residual_inventory_gate(available_inventory, total_capacity):
    return _safe_division(
        available_inventory,
        available_inventory + Constant(ILMENITE_ACTIVE_GATE_FRACTION) * Max(total_capacity, Constant(MIN_SOLID_CAPACITY_MOL_PER_M3)),
    )


def _arrhenius_expression(preexponential: float, activation_energy_j_per_mol: float, temperature_k):
    return Constant(preexponential) * Exp(-Constant(activation_energy_j_per_mol / GAS_CONSTANT_J_PER_MOL_K) / temperature_k)


def _gas_phase_concentration_expression(context: KineticsContext, species_id: str, temperature_k):
    gas_fraction = context.model.y_gas(context.gas_index(species_id), context.idx_cell)
    return _pressure_pa_expression(context.model.P(context.idx_cell)) * _positive_part_expression(gas_fraction) / (
        Constant(GAS_CONSTANT_J_PER_MOL_K) * temperature_k
    )


def _optional_gas_phase_concentration_expression(context: KineticsContext, species_id: str, temperature_k):
    try:
        return _gas_phase_concentration_expression(context, species_id, temperature_k)
    except KeyError:
        return Constant(0.0)


def _partial_pressure_kpa_expression(context: KineticsContext, species_id: str):
    gas_fraction = context.model.y_gas(context.gas_index(species_id), context.idx_cell)
    return _pressure_kpa_expression(context.model.P(context.idx_cell)) * _positive_part_expression(gas_fraction)


def _optional_partial_pressure_kpa_expression(context: KineticsContext, species_id: str):
    try:
        return _partial_pressure_kpa_expression(context, species_id)
    except KeyError:
        return Constant(0.0)


def _solid_concentration_expression(context: KineticsContext, species_id: str):
    concentration = context.model.c_sol(context.solid_index(species_id), context.idx_cell) / Constant(1.0 * mol / m**3)
    return _positive_part_expression(concentration)


def _optional_solid_concentration_expression(context: KineticsContext, species_id: str):
    try:
        return _solid_concentration_expression(context, species_id)
    except KeyError:
        return Constant(0.0)


def _oxygen_capacity_expression(c_fe2tio5, c_fe2tio4, c_fe2o3, c_feo):
    return c_fe2tio5 + c_fe2tio4 + c_fe2o3 + Constant(0.5) * c_feo


def _oxidized_inventory_expression(c_fe2tio5, c_fe2o3):
    return c_fe2tio5 + c_fe2o3


def _oxidation_demand_expression(c_fe2tio4, c_feo):
    return c_fe2tio4 + Constant(0.5) * c_feo


def _reaction_constant_expression(rate_key: str, temperature_k):
    return _arrhenius_expression(
        ILMENITE_RATE_COEFFICIENTS_M_PER_S[rate_key],
        ILMENITE_ACTIVATION_ENERGIES_J_PER_MOL[rate_key],
        temperature_k,
    )


def _diffusivity_expression(rate_key: str, temperature_k, conversion):
    return _arrhenius_expression(
        ILMENITE_DIFFUSIVITY_COEFFICIENTS_M2_PER_S[rate_key],
        ILMENITE_DIFFUSION_ACTIVATION_ENERGIES_J_PER_MOL[rate_key],
        temperature_k,
    ) * Exp(-Constant(ILMENITE_CONVERSION_DECAY[rate_key]) * conversion)


def _mixed_control_conversion_rate_expression(rate_key: str, temperature_k, gas_concentration_mol_per_m3, conversion):
    conversion_bounded = _bounded_conversion_expression(conversion)
    one_minus_x = Max(Constant(1.0) - conversion_bounded, Constant(MIN_ONE_MINUS_CONVERSION))
    # Keep the symbolic rate exactly zero when the reacting gas is absent.
    # Using a positive floor here makes the startup state inconsistent for
    # staged runs that begin under inert gas, even though the corresponding
    # value-level helper already uses max(c_g, 0.0).
    gas_concentration_safe = Max(gas_concentration_mol_per_m3, Constant(0.0))
    denominator = (
        Constant(1.0) / _reaction_constant_expression(rate_key, temperature_k) * one_minus_x ** Constant(-2.0 / 3.0)
        + Constant(ILMENITE_PARTICLE_RADIUS_M) / _diffusivity_expression(rate_key, temperature_k, conversion_bounded)
        * (one_minus_x ** Constant(-1.0 / 3.0) - Constant(1.0))
    )
    return (
        Constant(
            3.0
            * ILMENITE_STOICHIOMETRIC_FACTOR[rate_key]
            / (ILMENITE_PARTICLE_RADIUS_M * ILMENITE_SOLID_CONCENTRATIONS_MOL_PER_M3[rate_key])
        )
        * gas_concentration_safe ** Constant(ILMENITE_REACTION_ORDER[rate_key])
        / denominator
    )


def _surface_control_specific_rate_expression(rate_key: str, temperature_k, gas_concentration_mol_per_m3, conversion):
    conversion_bounded = _bounded_conversion_expression(conversion)
    gas_concentration_safe = Max(gas_concentration_mol_per_m3, Constant(0.0))
    return (
        Constant(
            3.0
            * ILMENITE_STOICHIOMETRIC_FACTOR[rate_key]
            / (ILMENITE_PARTICLE_RADIUS_M * ILMENITE_SOLID_CONCENTRATIONS_MOL_PER_M3[rate_key])
        )
        * _reaction_constant_expression(rate_key, temperature_k)
        * gas_concentration_safe ** Constant(ILMENITE_REACTION_ORDER[rate_key])
        * Exp(-Constant(ILMENITE_CONVERSION_DECAY[rate_key]) * conversion_bounded)
    )


def _ilmenite_catalyst_mass_density_expression(context: KineticsContext):
    return (
        _optional_solid_concentration_expression(context, "Fe2TiO5") * Constant(FE2TIO5_MW_KG_PER_MOL)
        + _optional_solid_concentration_expression(context, "Fe2TiO4") * Constant(FE2TIO4_MW_KG_PER_MOL)
        + _optional_solid_concentration_expression(context, "Fe2O3") * Constant(FE2O3_MW_KG_PER_MOL)
        + _optional_solid_concentration_expression(context, "FeO") * Constant(FEO_MW_KG_PER_MOL)
        + _optional_solid_concentration_expression(context, "Fe3O4") * Constant(FE3O4_MW_KG_PER_MOL)
        + _optional_solid_concentration_expression(context, "TiO2") * Constant(TIO2_MW_KG_PER_MOL)
    )


def _ilmenite_terms(context: KineticsContext) -> IlmeniteRedoxTerms:
    temperature_k = _temperature_k_expression(context.model.T(context.idx_cell))
    c_fe2tio5 = _optional_solid_concentration_expression(context, "Fe2TiO5")
    c_fe2tio4 = _optional_solid_concentration_expression(context, "Fe2TiO4")
    c_fe2o3 = _optional_solid_concentration_expression(context, "Fe2O3")
    c_feo = _optional_solid_concentration_expression(context, "FeO")
    oxygen_capacity = _oxygen_capacity_expression(c_fe2tio5, c_fe2tio4, c_fe2o3, c_feo)
    oxidized_inventory = _oxidized_inventory_expression(c_fe2tio5, c_fe2o3)
    oxidation_demand = _oxidation_demand_expression(c_fe2tio4, c_feo)

    return IlmeniteRedoxTerms(
        temperature_k=temperature_k,
        pressure_kpa=_pressure_kpa_expression(context.model.P(context.idx_cell)),
        c_h2_mol_per_m3=_optional_gas_phase_concentration_expression(context, "H2", temperature_k),
        c_co_mol_per_m3=_optional_gas_phase_concentration_expression(context, "CO", temperature_k),
        c_o2_mol_per_m3=_optional_gas_phase_concentration_expression(context, "O2", temperature_k),
        p_h2_kpa=_optional_partial_pressure_kpa_expression(context, "H2"),
        p_h2o_kpa=_optional_partial_pressure_kpa_expression(context, "H2O"),
        p_co_kpa=_optional_partial_pressure_kpa_expression(context, "CO"),
        p_co2_kpa=_optional_partial_pressure_kpa_expression(context, "CO2"),
        c_fe2tio5_mol_per_m3=c_fe2tio5,
        c_fe2tio4_mol_per_m3=c_fe2tio4,
        c_fe2o3_mol_per_m3=c_fe2o3,
        c_feo_mol_per_m3=c_feo,
        oxygen_capacity_mol_per_m3=oxygen_capacity,
        oxidized_inventory_mol_per_m3=oxidized_inventory,
        oxidation_demand_mol_per_m3=oxidation_demand,
        x_reduction=Constant(1.0) - _safe_division(oxidized_inventory, oxygen_capacity),
        x_oxidation=_safe_division(oxidized_inventory, oxygen_capacity),
        catalyst_mass_density_kg_per_m3=_ilmenite_catalyst_mass_density_expression(context),
    )


@register_kinetics_hook("ilmenite_fe2tio5_h2_reduction")
def ilmenite_fe2tio5_h2_reduction(context: KineticsContext):
    terms = _ilmenite_terms(context)
    total_site_rate = terms.oxygen_capacity_mol_per_m3 * _mixed_control_conversion_rate_expression(
        "h2_reduction",
        terms.temperature_k,
        terms.c_h2_mol_per_m3,
        terms.x_reduction,
    )
    share = _safe_division(terms.c_fe2tio5_mol_per_m3, terms.oxidized_inventory_mol_per_m3)
    return Constant(1.0 * mol / (m**3 * s)) * total_site_rate * share


@register_kinetics_hook("ilmenite_fe2o3_h2_reduction")
def ilmenite_fe2o3_h2_reduction(context: KineticsContext):
    terms = _ilmenite_terms(context)
    total_site_rate = terms.oxygen_capacity_mol_per_m3 * _mixed_control_conversion_rate_expression(
        "h2_reduction",
        terms.temperature_k,
        terms.c_h2_mol_per_m3,
        terms.x_reduction,
    )
    share = _safe_division(terms.c_fe2o3_mol_per_m3, terms.oxidized_inventory_mol_per_m3)
    return Constant(1.0 * mol / (m**3 * s)) * total_site_rate * share


@register_kinetics_hook("ilmenite_fe2tio5_co_reduction")
def ilmenite_fe2tio5_co_reduction(context: KineticsContext):
    terms = _ilmenite_terms(context)
    total_site_rate = terms.oxygen_capacity_mol_per_m3 * _mixed_control_conversion_rate_expression(
        "co_reduction",
        terms.temperature_k,
        terms.c_co_mol_per_m3,
        terms.x_reduction,
    )
    share = _safe_division(terms.c_fe2tio5_mol_per_m3, terms.oxidized_inventory_mol_per_m3)
    return Constant(1.0 * mol / (m**3 * s)) * total_site_rate * share


@register_kinetics_hook("ilmenite_fe2o3_co_reduction")
def ilmenite_fe2o3_co_reduction(context: KineticsContext):
    terms = _ilmenite_terms(context)
    total_site_rate = terms.oxygen_capacity_mol_per_m3 * _mixed_control_conversion_rate_expression(
        "co_reduction",
        terms.temperature_k,
        terms.c_co_mol_per_m3,
        terms.x_reduction,
    )
    share = _safe_division(terms.c_fe2o3_mol_per_m3, terms.oxidized_inventory_mol_per_m3)
    return Constant(1.0 * mol / (m**3 * s)) * total_site_rate * share


@register_kinetics_hook("ilmenite_fe2tio4_o2_oxidation")
def ilmenite_fe2tio4_o2_oxidation(context: KineticsContext):
    terms = _ilmenite_terms(context)
    total_site_rate = terms.oxygen_capacity_mol_per_m3 * _mixed_control_conversion_rate_expression(
        "o2_oxidation",
        terms.temperature_k,
        terms.c_o2_mol_per_m3,
        terms.x_oxidation,
    )
    share = _safe_division(terms.c_fe2tio4_mol_per_m3, terms.oxidation_demand_mol_per_m3)
    return Constant(1.0 * mol / (m**3 * s)) * total_site_rate * share


@register_kinetics_hook("ilmenite_feo_o2_oxidation")
def ilmenite_feo_o2_oxidation(context: KineticsContext):
    terms = _ilmenite_terms(context)
    total_site_rate = terms.oxygen_capacity_mol_per_m3 * _mixed_control_conversion_rate_expression(
        "o2_oxidation",
        terms.temperature_k,
        terms.c_o2_mol_per_m3,
        terms.x_oxidation,
    )
    share = _safe_division(Constant(0.5) * terms.c_feo_mol_per_m3, terms.oxidation_demand_mol_per_m3)
    return Constant(1.0 * mol / (m**3 * s)) * total_site_rate * share


@register_kinetics_hook("ilmenite_fe2tiox_h2_reduction_proxy")
def ilmenite_fe2tiox_h2_reduction_proxy(context: KineticsContext):
    temperature_k = _temperature_k_expression(context.model.T(context.idx_cell))
    c_fe2tio5 = _solid_concentration_expression(context, "Fe2TiO5")
    c_fe2tio4 = _solid_concentration_expression(context, "Fe2TiO4")
    oxygen_capacity = c_fe2tio5 + c_fe2tio4
    total_site_rate = c_fe2tio5 * _surface_control_specific_rate_expression(
        "h2_reduction",
        temperature_k,
        _gas_phase_concentration_expression(context, "H2", temperature_k),
        Constant(1.0) - _safe_division(c_fe2tio5, oxygen_capacity),
    )
    return Constant(1.0 * mol / (m**3 * s)) * total_site_rate


@register_kinetics_hook("ilmenite_fe2tiox_o2_oxidation_proxy")
def ilmenite_fe2tiox_o2_oxidation_proxy(context: KineticsContext):
    temperature_k = _temperature_k_expression(context.model.T(context.idx_cell))
    c_fe2tio5 = _solid_concentration_expression(context, "Fe2TiO5")
    c_fe2tio4 = _solid_concentration_expression(context, "Fe2TiO4")
    oxygen_capacity = c_fe2tio5 + c_fe2tio4
    total_site_rate = c_fe2tio4 * _surface_control_specific_rate_expression(
        "o2_oxidation",
        temperature_k,
        _gas_phase_concentration_expression(context, "O2", temperature_k),
        _safe_division(c_fe2tio5, oxygen_capacity),
    )
    return Constant(1.0 * mol / (m**3 * s)) * total_site_rate


def _equilibrium_constant_wgs_expression(temperature_k):
    z_value = Constant(1000.0) / temperature_k - Constant(1.0)
    return Exp(z_value * (z_value * (Constant(0.63508) - Constant(0.29353) * z_value) + Constant(4.1778)) + Constant(0.31688))


@register_kinetics_hook("ilmenite_wgs")
def ilmenite_wgs(context: KineticsContext):
    terms = _ilmenite_terms(context)

    p_co_available = Max(terms.p_co_kpa, Constant(0.0))
    p_h2o_available = Max(terms.p_h2o_kpa, Constant(0.0))
    p_co_safe = Max(terms.p_co_kpa, Constant(MIN_PARTIAL_PRESSURE_KPA))
    p_h2o_safe = Max(terms.p_h2o_kpa, Constant(MIN_PARTIAL_PRESSURE_KPA))
    p_co2_safe = Max(terms.p_co2_kpa, Constant(MIN_PARTIAL_PRESSURE_KPA))
    p_h2_safe = Max(terms.p_h2_kpa, Constant(MIN_PARTIAL_PRESSURE_KPA))
    reactant_gate = (
        p_co_available / (p_co_available + Constant(WGS_REACTANT_GATE_KPA))
        * p_h2o_available / (p_h2o_available + Constant(WGS_REACTANT_GATE_KPA))
    )
    beta = p_co2_safe * p_h2_safe / (_equilibrium_constant_wgs_expression(terms.temperature_k) * p_co_safe * p_h2o_safe)
    rate_expression = (
        reactant_gate
        * Constant(1.0e3)
        * terms.catalyst_mass_density_kg_per_m3
        * _arrhenius_expression(
            ILMENITE_WGS_PREEXPONENTIAL_MOL_PER_G_S,
            ILMENITE_WGS_ACTIVATION_ENERGY_J_PER_MOL,
            terms.temperature_k,
        )
        * p_co_safe ** Constant(ILMENITE_WGS_REACTION_ORDERS["CO"])
        * p_h2o_safe ** Constant(ILMENITE_WGS_REACTION_ORDERS["H2O"])
        * p_co2_safe ** Constant(ILMENITE_WGS_REACTION_ORDERS["CO2"])
        * p_h2_safe ** Constant(ILMENITE_WGS_REACTION_ORDERS["H2"])
        * (Constant(1.0) - beta)
    )
    return Constant(1.0 * mol / (m**3 * s)) * rate_expression


PAPER_G1_SAMPLE_MASS_OXIDIZED_KG = 100.0e-6
PAPER_G1_NOMINAL_ILMENITE_MASS_FRACTION = 0.85
PAPER_G1_NOMINAL_TIO2_MASS_FRACTION = 0.15
PAPER_G1_BULK_DENSITY_KG_PER_M3 = 1.29e3
PAPER_G1_GRAIN_POROSITY = 0.314
PAPER_G1_PARTICLE_DIAMETER_M = 3.6e-3
PAPER_TGA_TUBE_RADIUS_M = 7.5e-3
PAPER_TGA_INTERPARTICLE_VOIDAGE = 0.40
# The TGA kinetics runs are reported at 840 mL/min (STP), not 8 mL/s.
PAPER_TGA_TOTAL_FLOW_MOL_PER_S = 840.0e-6 / 60.0 / 22.414e-3
PAPER_TGA_PRESSURE_PA = 101325.0

PAPER_G1_NOMINAL_ILMENITE_MOLES = PAPER_G1_NOMINAL_ILMENITE_MASS_FRACTION / FETIO3_MW_KG_PER_MOL
PAPER_G1_BASE_TIO2_MOLES = PAPER_G1_NOMINAL_ILMENITE_MOLES + PAPER_G1_NOMINAL_TIO2_MASS_FRACTION / TIO2_MW_KG_PER_MOL
PAPER_G1_BASE_FE2O3_MOLES = 0.5 * PAPER_G1_NOMINAL_ILMENITE_MOLES
PAPER_G1_BASE_FEO_MOLES = PAPER_G1_NOMINAL_ILMENITE_MOLES
PAPER_G1_BASE_OXIDIZED_MASS_KG = (
    PAPER_G1_BASE_FE2O3_MOLES * FE2O3_MW_KG_PER_MOL
    + PAPER_G1_BASE_TIO2_MOLES * TIO2_MW_KG_PER_MOL
)
PAPER_G1_OXYGEN_INVENTORY_SCALE = PAPER_G1_SAMPLE_MASS_OXIDIZED_KG / PAPER_G1_BASE_OXIDIZED_MASS_KG
PAPER_TGA_BED_SOLID_VOLUME_M3 = PAPER_G1_SAMPLE_MASS_OXIDIZED_KG / PAPER_G1_BULK_DENSITY_KG_PER_M3
PAPER_TGA_BED_VOLUME_M3 = PAPER_TGA_BED_SOLID_VOLUME_M3 / (1.0 - PAPER_TGA_INTERPARTICLE_VOIDAGE)
PAPER_TGA_BED_LENGTH_M = PAPER_TGA_BED_VOLUME_M3 / (math.pi * PAPER_TGA_TUBE_RADIUS_M**2)


def paper_tga_bed_volume_m3() -> float:
    return PAPER_TGA_BED_VOLUME_M3


def paper_tga_bed_length_m() -> float:
    return PAPER_TGA_BED_LENGTH_M


def paper_oxidized_bed_profile_mol_per_m3() -> dict[str, float]:
    fe2o3_moles = PAPER_G1_OXYGEN_INVENTORY_SCALE * PAPER_G1_BASE_FE2O3_MOLES
    tio2_moles = PAPER_G1_OXYGEN_INVENTORY_SCALE * PAPER_G1_BASE_TIO2_MOLES
    return {
        "Fe2O3": fe2o3_moles / PAPER_TGA_BED_VOLUME_M3,
        "FeO": 0.0,
        "TiO2": tio2_moles / PAPER_TGA_BED_VOLUME_M3,
    }


def paper_reduced_bed_profile_mol_per_m3() -> dict[str, float]:
    feo_moles = PAPER_G1_OXYGEN_INVENTORY_SCALE * PAPER_G1_BASE_FEO_MOLES
    tio2_moles = PAPER_G1_OXYGEN_INVENTORY_SCALE * PAPER_G1_BASE_TIO2_MOLES
    return {
        "Fe2O3": 0.0,
        "FeO": feo_moles / PAPER_TGA_BED_VOLUME_M3,
        "TiO2": tio2_moles / PAPER_TGA_BED_VOLUME_M3,
    }


def _write_mapping_as_yaml_json(path: Path, mapping: Mapping[str, Any]) -> None:
    import yaml

    class _PlainFloatDumper(yaml.SafeDumper):
        pass

    def _represent_float(dumper, value: float):
        text = format(value, ".15f").rstrip("0").rstrip(".")
        if "." not in text:
            text = f"{text}.0"
        return dumper.represent_scalar("tag:yaml.org,2002:float", text)

    _PlainFloatDumper.add_representer(float, _represent_float)
    serialized = yaml.dump(dict(mapping), Dumper=_PlainFloatDumper, sort_keys=False)
    path.write_text(serialized, encoding="utf-8")


def _paper_validation_run_config(
    *,
    horizon_s: float,
    reporting_interval_s: float,
    system_name: str,
    wall_temperature_k: float,
) -> dict[str, Any]:
    return {
        "references": {
            "chemistry_file": "chemistry.yaml",
            "program_file": "program.yaml",
            "solids_file": "solids.yaml",
        },
        "simulation": {
            "system_name": system_name,
            "time_horizon_s": horizon_s,
            "reporting_interval_s": reporting_interval_s,
            "repeat_program": False,
            "mass_scheme": "upwind1",
            "heat_scheme": "upwind1",
            "report_time_derivatives": False,
        },
        "model": {
            "bed_length_m": PAPER_TGA_BED_LENGTH_M,
            "bed_radius_m": PAPER_TGA_TUBE_RADIUS_M,
            "axial_cells": 3,
            "isothermal_bed": True,
            "wall_temperature_k": wall_temperature_k,
        },
        "solver": {
            "relative_tolerance": 0.000001,
        },
        "outputs": {
            "directory": "output",
            "artifacts_directory": "output/artifacts",
            "requested_reports": ["solid_concentration"],
        },
    }


def _paper_validation_chemistry_config(*, gas_species: tuple[str, ...], reaction_id: str) -> dict[str, Any]:
    return {
        "gas_species": list(gas_species),
        "reaction_ids": [reaction_id],
    }


def _paper_validation_program_config(*, temperature_k: float, composition: Mapping[str, float]) -> dict[str, Any]:
    return {
        "inlet_flow": {
            "initial": PAPER_TGA_TOTAL_FLOW_MOL_PER_S,
        },
        "inlet_temperature": {
            "initial": temperature_k,
        },
        "outlet_pressure": {
            "initial": PAPER_TGA_PRESSURE_PA,
        },
        "inlet_composition": {
            "initial": dict(composition),
        },
    }


def _paper_validation_solids_config(*, profile_values: Mapping[str, float]) -> dict[str, Any]:
    return {
        "solid_species": ["Fe2O3", "FeO", "TiO2"],
        "initial_profile": {
            "basis": "bed",
            "zones": [
                {
                    "x_start_m": 0.0,
                    "x_end_m": PAPER_TGA_BED_LENGTH_M,
                    "e_b": PAPER_TGA_INTERPARTICLE_VOIDAGE,
                    "e_p": PAPER_G1_GRAIN_POROSITY,
                    "d_p": PAPER_G1_PARTICLE_DIAMETER_M,
                    "values": dict(profile_values),
                }
            ],
        },
    }


def _find_process_variable(process, variable_name: str):
    matches = sorted(
        key
        for key in getattr(process, "dictVariables", {})
        if key == variable_name or key.endswith(f".{variable_name}")
    )
    if len(matches) != 1:
        return None
    return process.dictVariables[matches[0]]


def _collapse_duplicate_times(time_s, values):
    import numpy as np

    time_array = np.asarray(time_s, dtype=float).reshape(-1)
    value_array = np.asarray(values, dtype=float)
    if time_array.size <= 1:
        return time_array, value_array
    _, first_indices = np.unique(time_array, return_index=True)
    first_indices = np.sort(first_indices)
    return time_array[first_indices], value_array[first_indices]


def _extract_conversion_trace_from_reporter(reporter, *, bed_length_m: float, bed_radius_m: float, species_order: tuple[str, ...], phase: str):
    import numpy as np

    process = getattr(reporter, "Process", None)
    if process is None or not hasattr(process, "dictVariables"):
        raise RuntimeError("Reporter does not expose the process variables needed for ilmenite validation.")

    solid_variable = _find_process_variable(process, "c_sol")
    if solid_variable is None:
        raise RuntimeError("solid_concentration report was not available in the ilmenite validation run.")

    time_s, solid_concentration = _collapse_duplicate_times(solid_variable.TimeValues, solid_variable.Values)
    if solid_concentration.ndim != 3:
        raise RuntimeError("Unexpected solid concentration shape in ilmenite validation reporter output.")

    area_m2 = math.pi * bed_radius_m**2
    cell_width_m = bed_length_m / solid_concentration.shape[2]
    species_inventory_mol = area_m2 * cell_width_m * solid_concentration.sum(axis=2)

    if phase == "reduction":
        species_id = "Fe2O3"
    elif phase == "oxidation":
        species_id = "FeO"
    else:
        raise ValueError(f"Unsupported ilmenite validation phase '{phase}'.")

    species_index = species_order.index(species_id)
    initial_inventory = float(species_inventory_mol[0, species_index])
    if initial_inventory <= 0.0:
        raise RuntimeError(f"Initial {species_id} inventory was zero in ilmenite validation run.")

    conversion = 1.0 - species_inventory_mol[:, species_index] / initial_inventory
    return np.asarray(time_s, dtype=float), np.clip(np.asarray(conversion, dtype=float), 0.0, 1.0)


def _run_paper_validation_case(
    *,
    system_name: str,
    horizon_s: float,
    reporting_interval_s: float,
    gas_species: tuple[str, ...],
    reaction_id: str,
    temperature_k: float,
    composition: Mapping[str, float],
    initial_profile: Mapping[str, float],
    phase: str,
):
    import contextlib
    import io

    from ..config import load_run_bundle
    from ..properties import PROPERTY_REGISTRY as default_property_registry
    from ..reactions import REACTION_CATALOG as default_reaction_catalog
    from ..solver_clean import assemble_simulation, run_assembled_simulation

    temporary_root = Path(__file__).resolve().parents[2] / ".ilmenite_validation_tmp"
    temporary_dir = temporary_root / system_name
    if temporary_dir.exists():
        shutil.rmtree(temporary_dir, ignore_errors=True)
    temporary_dir.mkdir(parents=True, exist_ok=True)
    try:
        run_yaml_path = temporary_dir / "run.yaml"
        _write_mapping_as_yaml_json(
            run_yaml_path,
            _paper_validation_run_config(
                horizon_s=horizon_s,
                reporting_interval_s=reporting_interval_s,
                system_name=system_name,
                wall_temperature_k=temperature_k,
            ),
        )
        _write_mapping_as_yaml_json(
            temporary_dir / "chemistry.yaml",
            _paper_validation_chemistry_config(gas_species=gas_species, reaction_id=reaction_id),
        )
        _write_mapping_as_yaml_json(
            temporary_dir / "program.yaml",
            _paper_validation_program_config(temperature_k=temperature_k, composition=composition),
        )
        _write_mapping_as_yaml_json(
            temporary_dir / "solids.yaml",
            _paper_validation_solids_config(profile_values=initial_profile),
        )

        run_bundle = load_run_bundle(run_yaml_path)
        assembly = assemble_simulation(
            run_bundle,
            property_registry=default_property_registry,
            reaction_catalog=default_reaction_catalog,
        )
        with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
            reporter = run_assembled_simulation(assembly, include_plot_variables=False)
        return _extract_conversion_trace_from_reporter(
            reporter,
            bed_length_m=run_bundle.run.model.bed_length_m,
            bed_radius_m=run_bundle.run.model.bed_radius_m,
            species_order=run_bundle.solids.solid_species,
            phase=phase,
        )
    finally:
        shutil.rmtree(temporary_dir, ignore_errors=True)


def _plot_paper_validation_panel(
    axis,
    *,
    panel_title: str,
    traces: list[tuple[str, Any, Any]],
    paper_traces: Mapping[str, tuple[Any, Any]] | None,
    x_limit_s: float,
    y_label: str,
):
    color_cycle = ("#0c4a6e", "#b91c1c", "#166534", "#7c3aed", "#c2410c")
    for index, (label, time_s, conversion) in enumerate(traces):
        color = color_cycle[index % len(color_cycle)]
        axis.plot(
            time_s,
            conversion,
            linewidth=2.0,
            color=color,
            label=label,
        )
        if paper_traces is not None and label in paper_traces:
            paper_time_s, paper_conversion = paper_traces[label]
            axis.plot(
                paper_time_s,
                paper_conversion,
                linewidth=1.8,
                linestyle="--",
                color=color,
            )
    axis.set_xlim(0.0, x_limit_s)
    axis.set_ylim(0.0, 1.02)
    axis.set_xlabel("Time [s]")
    axis.set_ylabel(y_label)
    axis.set_title(panel_title)
    axis.grid(True, which="major", color="#d4d4d8", linewidth=0.7)
    axis.text(
        0.03,
        0.97,
        "solid = code\ndashed = experimental data",
        transform=axis.transAxes,
        ha="left",
        va="top",
        fontsize=9,
        bbox={"facecolor": "white", "edgecolor": "#d4d4d8", "alpha": 0.9, "boxstyle": "round,pad=0.25"},
    )
    axis.legend(frameon=False, loc="lower right")


def _normalize_paper_trace_label(raw_label: Any) -> str | None:
    text = str(raw_label).strip()
    if not text:
        return None
    text = " ".join(text.replace("°", "").split())
    if "%" in text:
        return text
    try:
        numeric_value = float(text)
    except ValueError:
        return text
    if numeric_value.is_integer():
        return f"{int(numeric_value)} C"
    return f"{numeric_value:g} C"


def _load_digitized_paper_traces(panel_name: str) -> dict[str, tuple[Any, Any]]:
    import numpy as np

    workbook_names = {
        "figure_2a": "2a.xlsx",
        "figure_2b": "2b.xlsx",
        "figure_3a": "3a.xlsx",
        "figure_3b": "3b.xlsx",
        "figure_4a": "4a.xlsx",
        "figure_4b": "4b.xlsx",
    }
    workbook_path = Path(__file__).resolve().parents[2] / workbook_names[panel_name]
    if workbook_path.exists():
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        traces: dict[str, tuple[Any, Any]] = {}
        if not rows:
            return traces

        header_row = rows[0]
        for column_index, header_value in enumerate(header_row):
            label = _normalize_paper_trace_label(header_value)
            if label is None:
                continue
            points: list[tuple[float, float]] = []
            for row in rows[2:]:
                if column_index + 1 >= len(row):
                    continue
                time_value = row[column_index]
                conversion_value = row[column_index + 1]
                if time_value is None or conversion_value is None:
                    continue
                points.append((float(time_value), float(conversion_value)))
            if not points:
                continue
            points.sort(key=lambda item: item[0])
            time_s = np.asarray([item[0] for item in points], dtype=float)
            x_plot = np.asarray([item[1] for item in points], dtype=float)
            traces[label] = (time_s, np.clip(x_plot, 0.0, 1.0))
        return traces

    import csv

    path = (
        Path(__file__).resolve().parents[1]
        / "examples"
        / "ilmenite_redox_case"
        / "digitized"
        / f"{panel_name}.csv"
    )
    grouped: dict[str, list[tuple[float, float]]] = {}
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            label = row["label"].strip()
            grouped.setdefault(label, []).append((float(row["time_s"]), float(row["x_plot"])))

    traces: dict[str, tuple[Any, Any]] = {}
    for label, rows in grouped.items():
        rows.sort(key=lambda item: item[0])
        time_s = np.asarray([item[0] for item in rows], dtype=float)
        x_plot = np.asarray([item[1] for item in rows], dtype=float)
        traces[label] = (time_s, np.clip(x_plot, 0.0, 1.0))
    return traces


def _paper_reference_trace(
    *,
    rate_key: str,
    temperature_k: float,
    gas_mole_fraction: float,
    horizon_s: float,
    reporting_interval_s: float,
):
    import numpy as np

    gas_concentration_mol_per_m3 = gas_phase_concentration_value(
        total_pressure_pa=PAPER_TGA_PRESSURE_PA,
        mole_fraction=gas_mole_fraction,
        temperature_k=temperature_k,
    )

    time_s = np.arange(0.0, horizon_s + 0.5 * reporting_interval_s, reporting_interval_s, dtype=float)
    conversion = np.zeros_like(time_s)
    max_internal_step_s = 0.05

    def rate(current_conversion: float) -> float:
        return mixed_control_conversion_rate_value(
            rate_key,
            temperature_k=temperature_k,
            gas_concentration_mol_per_m3=gas_concentration_mol_per_m3,
            conversion=current_conversion,
        )

    current_conversion = 0.0
    for index in range(1, time_s.size):
        remaining = float(time_s[index] - time_s[index - 1])
        while remaining > 1.0e-12:
            step_s = min(max_internal_step_s, remaining)
            k1 = rate(current_conversion)
            k2 = rate(min(max(current_conversion + 0.5 * step_s * k1, 0.0), 1.0))
            k3 = rate(min(max(current_conversion + 0.5 * step_s * k2, 0.0), 1.0))
            k4 = rate(min(max(current_conversion + step_s * k3, 0.0), 1.0))
            current_conversion = min(
                max(current_conversion + step_s * (k1 + 2.0 * k2 + 2.0 * k3 + k4) / 6.0, 0.0),
                1.0,
            )
            remaining -= step_s
        conversion[index] = current_conversion

    return time_s, conversion


def _rmse_between_traces(
    *,
    model_time_s,
    model_conversion,
    paper_time_s,
    paper_conversion,
) -> float:
    import numpy as np

    model_time = np.asarray(model_time_s, dtype=float)
    model_values = np.asarray(model_conversion, dtype=float)
    paper_time = np.asarray(paper_time_s, dtype=float)
    paper_values = np.asarray(paper_conversion, dtype=float)
    paper_interp = np.interp(model_time, paper_time, paper_values)
    return float(np.sqrt(np.mean((model_values - paper_interp) ** 2)))


def _legacy_run_ortiz_2016_tga_validation(output_dir: str | Path | None = None) -> dict[str, Path]:
    import matplotlib
    import numpy as np

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    if output_dir is None:
        output_root = Path(__file__).resolve().parents[1] / "examples" / "ilmenite_redox_case" / "output" / "artifacts"
    else:
        output_root = Path(output_dir).resolve()
    output_root.mkdir(parents=True, exist_ok=True)

    oxidized_profile = paper_oxidized_bed_profile_mol_per_m3()
    reduced_profile = paper_reduced_bed_profile_mol_per_m3()

    fig2a_traces = []
    for label, temperature_k in (
        ("700 °C", 973.15),
        ("800 °C", 1073.15),
        ("900 °C", 1173.15),
        ("1100 °C", 1373.15),
    ):
        time_s, conversion = _run_paper_validation_case(
            system_name=f"IlmeniteFig2A_{label.replace(' ', '').replace('°', '')}",
            horizon_s=200.0,
            reporting_interval_s=1.0,
            gas_species=("H2", "H2O", "N2"),
            reaction_id="ilmenite_fe2o3_h2_reduction_ortiz_2016",
            temperature_k=temperature_k,
            composition={"H2": 0.15, "H2O": 0.0, "N2": 0.85},
            initial_profile=oxidized_profile,
            phase="reduction",
        )
        fig2a_traces.append((label, time_s, conversion))

    fig2b_traces = []
    for label, h2_fraction in (
        ("15% H2", 0.15),
        ("30% H2", 0.30),
        ("50% H2", 0.50),
    ):
        time_s, conversion = _run_paper_validation_case(
            system_name=f"IlmeniteFig2B_{label.replace('%', '').replace(' ', '')}",
            horizon_s=200.0,
            reporting_interval_s=1.0,
            gas_species=("H2", "H2O", "N2"),
            reaction_id="ilmenite_fe2o3_h2_reduction_ortiz_2016",
            temperature_k=1073.15,
            composition={"H2": h2_fraction, "H2O": 0.0, "N2": 1.0 - h2_fraction},
            initial_profile=oxidized_profile,
            phase="reduction",
        )
        fig2b_traces.append((label, time_s, conversion))

    fig3a_traces = []
    for label, temperature_k in (
        ("600 °C", 873.15),
        ("800 °C", 1073.15),
        ("1000 °C", 1273.15),
    ):
        time_s, conversion = _run_paper_validation_case(
            system_name=f"IlmeniteFig3A_{label.replace(' ', '').replace('°', '')}",
            horizon_s=500.0,
            reporting_interval_s=2.0,
            gas_species=("CO", "CO2", "N2"),
            reaction_id="ilmenite_fe2o3_co_reduction_ortiz_2016",
            temperature_k=temperature_k,
            composition={"CO": 0.15, "CO2": 0.0, "N2": 0.85},
            initial_profile=oxidized_profile,
            phase="reduction",
        )
        fig3a_traces.append((label, time_s, conversion))

    fig3b_traces = []
    for label, co_fraction in (
        ("15% CO", 0.15),
        ("30% CO", 0.30),
        ("50% CO", 0.50),
    ):
        time_s, conversion = _run_paper_validation_case(
            system_name=f"IlmeniteFig3B_{label.replace('%', '').replace(' ', '')}",
            horizon_s=500.0,
            reporting_interval_s=2.0,
            gas_species=("CO", "CO2", "N2"),
            reaction_id="ilmenite_fe2o3_co_reduction_ortiz_2016",
            temperature_k=1073.15,
            composition={"CO": co_fraction, "CO2": 0.0, "N2": 1.0 - co_fraction},
            initial_profile=oxidized_profile,
            phase="reduction",
        )
        fig3b_traces.append((label, time_s, conversion))

    fig4a_traces = []
    for label, o2_fraction in (
        ("21% O2", 0.21),
        ("15% O2", 0.15),
        ("10% O2", 0.10),
        ("5% O2", 0.05),
    ):
        time_s, conversion = _run_paper_validation_case(
            system_name=f"IlmeniteFig4A_{label.replace('%', '').replace(' ', '')}",
            horizon_s=150.0,
            reporting_interval_s=1.0,
            gas_species=("O2", "N2"),
            reaction_id="ilmenite_feo_o2_oxidation_ortiz_2016",
            temperature_k=1073.15,
            composition={"O2": o2_fraction, "N2": 1.0 - o2_fraction},
            initial_profile=reduced_profile,
            phase="oxidation",
        )
        fig4a_traces.append((label, time_s, conversion))

    fig4b_traces = []
    for label, temperature_k in (
        ("700 °C", 973.15),
        ("800 °C", 1073.15),
        ("900 °C", 1173.15),
        ("1100 °C", 1373.15),
    ):
        time_s, conversion = _run_paper_validation_case(
            system_name=f"IlmeniteFig4B_{label.replace(' ', '').replace('°', '')}",
            horizon_s=80.0,
            reporting_interval_s=1.0,
            gas_species=("O2", "N2"),
            reaction_id="ilmenite_feo_o2_oxidation_ortiz_2016",
            temperature_k=temperature_k,
            composition={"O2": 0.21, "N2": 0.79},
            initial_profile=reduced_profile,
            phase="oxidation",
        )
        fig4b_traces.append((label, time_s, conversion))

    artifact_paths: dict[str, Path] = {}

    figure_2, axes_2 = plt.subplots(1, 2, figsize=(11.5, 4.6), constrained_layout=True)
    _plot_paper_validation_panel(
        axes_2[0],
        panel_title="Figure 2(a): H2 reduction, 15 vol% H2",
        traces=fig2a_traces,
        x_limit_s=200.0,
        y_label="Xred [-]",
    )
    _plot_paper_validation_panel(
        axes_2[1],
        panel_title="Figure 2(b): H2 reduction, 800 °C",
        traces=fig2b_traces,
        x_limit_s=200.0,
        y_label="Xred [-]",
    )
    figure_2_path = output_root / "ortiz_2016_fig2_solver_replication.svg"
    figure_2.savefig(figure_2_path, bbox_inches="tight")
    plt.close(figure_2)
    artifact_paths["figure_2"] = figure_2_path

    figure_3, axes_3 = plt.subplots(1, 2, figsize=(11.5, 4.6), constrained_layout=True)
    _plot_paper_validation_panel(
        axes_3[0],
        panel_title="Figure 3(a): CO reduction, 15 vol% CO",
        traces=fig3a_traces,
        x_limit_s=500.0,
        y_label="Xred [-]",
    )
    _plot_paper_validation_panel(
        axes_3[1],
        panel_title="Figure 3(b): CO reduction, 800 °C",
        traces=fig3b_traces,
        x_limit_s=500.0,
        y_label="Xred [-]",
    )
    figure_3_path = output_root / "ortiz_2016_fig3_solver_replication.svg"
    figure_3.savefig(figure_3_path, bbox_inches="tight")
    plt.close(figure_3)
    artifact_paths["figure_3"] = figure_3_path

    figure_4, axes_4 = plt.subplots(1, 2, figsize=(11.5, 4.6), constrained_layout=True)
    _plot_paper_validation_panel(
        axes_4[0],
        panel_title="Figure 4(a): O2 oxidation, 800 °C",
        traces=fig4a_traces,
        x_limit_s=150.0,
        y_label="Xox [-]",
    )
    _plot_paper_validation_panel(
        axes_4[1],
        panel_title="Figure 4(b): O2 oxidation, 21 vol% O2",
        traces=fig4b_traces,
        x_limit_s=80.0,
        y_label="Xox [-]",
    )
    figure_4_path = output_root / "ortiz_2016_fig4_solver_replication.svg"
    figure_4.savefig(figure_4_path, bbox_inches="tight")
    plt.close(figure_4)
    artifact_paths["figure_4"] = figure_4_path

    summary_rows = []
    for figure_name, traces in (
        ("figure_2a", fig2a_traces),
        ("figure_2b", fig2b_traces),
        ("figure_3a", fig3a_traces),
        ("figure_3b", fig3b_traces),
        ("figure_4a", fig4a_traces),
        ("figure_4b", fig4b_traces),
    ):
        for label, time_s, conversion in traces:
            summary_rows.append(
                {
                    "panel": figure_name,
                    "label": label,
                    "t_end_s": float(np.asarray(time_s, dtype=float)[-1]),
                    "x_end": float(np.asarray(conversion, dtype=float)[-1]),
                }
            )
    summary_path = output_root / "ortiz_2016_validation_trace_summary.json"
    summary_path.write_text(json.dumps(summary_rows, indent=2), encoding="utf-8")
    artifact_paths["summary"] = summary_path

    return artifact_paths


def run_ortiz_2016_tga_validation(output_dir: str | Path | None = None) -> dict[str, Path]:
    import matplotlib
    import numpy as np

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    if output_dir is None:
        output_root = Path(__file__).resolve().parents[1] / "examples" / "ilmenite_redox_case" / "output" / "artifacts"
    else:
        output_root = Path(output_dir).resolve()
    output_root.mkdir(parents=True, exist_ok=True)

    oxidized_profile = paper_oxidized_bed_profile_mol_per_m3()
    reduced_profile = paper_reduced_bed_profile_mol_per_m3()

    fig2a_paper_traces = _load_digitized_paper_traces("figure_2a")
    fig2a_traces = []
    for label, temperature_k in (
        ("700 C", 973.15),
        ("800 C", 1073.15),
        ("900 C", 1173.15),
        ("1100 C", 1373.15),
    ):
        time_s, conversion = _run_paper_validation_case(
            system_name=f"IlmeniteFig2A_{label.replace(' ', '')}",
            horizon_s=200.0,
            reporting_interval_s=1.0,
            gas_species=("H2", "H2O", "N2"),
            reaction_id="ilmenite_fe2o3_h2_reduction_ortiz_2016",
            temperature_k=temperature_k,
            composition={"H2": 0.15, "H2O": 0.0, "N2": 0.85},
            initial_profile=oxidized_profile,
            phase="reduction",
        )
        fig2a_traces.append((label, time_s, conversion))

    fig2b_paper_traces = _load_digitized_paper_traces("figure_2b")
    fig2b_traces = []
    for label, h2_fraction in (
        ("15% H2", 0.15),
        ("30% H2", 0.30),
        ("50% H2", 0.50),
    ):
        time_s, conversion = _run_paper_validation_case(
            system_name=f"IlmeniteFig2B_{label.replace('%', '').replace(' ', '')}",
            horizon_s=200.0,
            reporting_interval_s=1.0,
            gas_species=("H2", "H2O", "N2"),
            reaction_id="ilmenite_fe2o3_h2_reduction_ortiz_2016",
            temperature_k=1073.15,
            composition={"H2": h2_fraction, "H2O": 0.0, "N2": 1.0 - h2_fraction},
            initial_profile=oxidized_profile,
            phase="reduction",
        )
        fig2b_traces.append((label, time_s, conversion))

    fig3a_paper_traces = _load_digitized_paper_traces("figure_3a")
    fig3a_traces = []
    for label, temperature_k in (
        ("600 C", 873.15),
        ("800 C", 1073.15),
        ("1000 C", 1273.15),
    ):
        time_s, conversion = _run_paper_validation_case(
            system_name=f"IlmeniteFig3A_{label.replace(' ', '')}",
            horizon_s=500.0,
            reporting_interval_s=2.0,
            gas_species=("CO", "CO2", "N2"),
            reaction_id="ilmenite_fe2o3_co_reduction_ortiz_2016",
            temperature_k=temperature_k,
            composition={"CO": 0.15, "CO2": 0.0, "N2": 0.85},
            initial_profile=oxidized_profile,
            phase="reduction",
        )
        fig3a_traces.append((label, time_s, conversion))

    fig3b_paper_traces = _load_digitized_paper_traces("figure_3b")
    fig3b_traces = []
    for label, co_fraction in (
        ("15% CO", 0.15),
        ("30% CO", 0.30),
        ("50% CO", 0.50),
    ):
        time_s, conversion = _run_paper_validation_case(
            system_name=f"IlmeniteFig3B_{label.replace('%', '').replace(' ', '')}",
            horizon_s=500.0,
            reporting_interval_s=2.0,
            gas_species=("CO", "CO2", "N2"),
            reaction_id="ilmenite_fe2o3_co_reduction_ortiz_2016",
            temperature_k=1073.15,
            composition={"CO": co_fraction, "CO2": 0.0, "N2": 1.0 - co_fraction},
            initial_profile=oxidized_profile,
            phase="reduction",
        )
        fig3b_traces.append((label, time_s, conversion))

    fig4a_paper_traces = _load_digitized_paper_traces("figure_4a")
    fig4a_traces = []
    for label, o2_fraction in (
        ("21% O2", 0.21),
        ("15% O2", 0.15),
        ("10% O2", 0.10),
        ("5% O2", 0.05),
    ):
        time_s, conversion = _run_paper_validation_case(
            system_name=f"IlmeniteFig4A_{label.replace('%', '').replace(' ', '')}",
            horizon_s=150.0,
            reporting_interval_s=1.0,
            gas_species=("O2", "N2"),
            reaction_id="ilmenite_feo_o2_oxidation_ortiz_2016",
            temperature_k=1073.15,
            composition={"O2": o2_fraction, "N2": 1.0 - o2_fraction},
            initial_profile=reduced_profile,
            phase="oxidation",
        )
        fig4a_traces.append((label, time_s, conversion))

    fig4b_paper_traces = _load_digitized_paper_traces("figure_4b")
    fig4b_traces = []
    for label, temperature_k in (
        ("700 C", 973.15),
        ("800 C", 1073.15),
        ("900 C", 1173.15),
        ("1100 C", 1373.15),
    ):
        time_s, conversion = _run_paper_validation_case(
            system_name=f"IlmeniteFig4B_{label.replace(' ', '')}",
            horizon_s=80.0,
            reporting_interval_s=1.0,
            gas_species=("O2", "N2"),
            reaction_id="ilmenite_feo_o2_oxidation_ortiz_2016",
            temperature_k=temperature_k,
            composition={"O2": 0.21, "N2": 0.79},
            initial_profile=reduced_profile,
            phase="oxidation",
        )
        fig4b_traces.append((label, time_s, conversion))

    artifact_paths: dict[str, Path] = {}

    figure_2, axes_2 = plt.subplots(1, 2, figsize=(11.5, 4.6), constrained_layout=True)
    _plot_paper_validation_panel(
        axes_2[0],
        panel_title="Figure 2(a): H2 reduction, 15 vol% H2",
        traces=fig2a_traces,
        paper_traces=fig2a_paper_traces,
        x_limit_s=200.0,
        y_label="Xred [-]",
    )
    _plot_paper_validation_panel(
        axes_2[1],
        panel_title="Figure 2(b): H2 reduction, 800 C",
        traces=fig2b_traces,
        paper_traces=fig2b_paper_traces,
        x_limit_s=200.0,
        y_label="Xred [-]",
    )
    figure_2_path = output_root / "ortiz_2016_fig2_solver_replication.svg"
    figure_2.savefig(figure_2_path, bbox_inches="tight")
    plt.close(figure_2)
    artifact_paths["figure_2"] = figure_2_path

    figure_3, axes_3 = plt.subplots(1, 2, figsize=(11.5, 4.6), constrained_layout=True)
    _plot_paper_validation_panel(
        axes_3[0],
        panel_title="Figure 3(a): CO reduction, 15 vol% CO",
        traces=fig3a_traces,
        paper_traces=fig3a_paper_traces,
        x_limit_s=500.0,
        y_label="Xred [-]",
    )
    _plot_paper_validation_panel(
        axes_3[1],
        panel_title="Figure 3(b): CO reduction, 800 C",
        traces=fig3b_traces,
        paper_traces=fig3b_paper_traces,
        x_limit_s=500.0,
        y_label="Xred [-]",
    )
    figure_3_path = output_root / "ortiz_2016_fig3_solver_replication.svg"
    figure_3.savefig(figure_3_path, bbox_inches="tight")
    plt.close(figure_3)
    artifact_paths["figure_3"] = figure_3_path

    figure_4, axes_4 = plt.subplots(1, 2, figsize=(11.5, 4.6), constrained_layout=True)
    _plot_paper_validation_panel(
        axes_4[0],
        panel_title="Figure 4(a): O2 oxidation, 800 C",
        traces=fig4a_traces,
        paper_traces=fig4a_paper_traces,
        x_limit_s=150.0,
        y_label="Xox [-]",
    )
    _plot_paper_validation_panel(
        axes_4[1],
        panel_title="Figure 4(b): O2 oxidation, 21 vol% O2",
        traces=fig4b_traces,
        paper_traces=fig4b_paper_traces,
        x_limit_s=80.0,
        y_label="Xox [-]",
    )
    figure_4_path = output_root / "ortiz_2016_fig4_solver_replication.svg"
    figure_4.savefig(figure_4_path, bbox_inches="tight")
    plt.close(figure_4)
    artifact_paths["figure_4"] = figure_4_path

    summary_rows = []
    for figure_name, traces, paper_traces in (
        ("figure_2a", fig2a_traces, fig2a_paper_traces),
        ("figure_2b", fig2b_traces, fig2b_paper_traces),
        ("figure_3a", fig3a_traces, fig3a_paper_traces),
        ("figure_3b", fig3b_traces, fig3b_paper_traces),
        ("figure_4a", fig4a_traces, fig4a_paper_traces),
        ("figure_4b", fig4b_traces, fig4b_paper_traces),
    ):
        for label, time_s, conversion in traces:
            paper_time_s, paper_conversion = paper_traces[label]
            summary_rows.append(
                {
                    "panel": figure_name,
                    "label": label,
                    "t_end_s": float(np.asarray(time_s, dtype=float)[-1]),
                    "x_end": float(np.asarray(conversion, dtype=float)[-1]),
                    "paper_x_end": float(np.asarray(paper_conversion, dtype=float)[-1]),
                    "paper_source": "paper_plot_workbook",
                    "rmse_vs_paper": _rmse_between_traces(
                        model_time_s=time_s,
                        model_conversion=conversion,
                        paper_time_s=paper_time_s,
                        paper_conversion=paper_conversion,
                    ),
                }
            )
    summary_path = output_root / "ortiz_2016_validation_trace_summary.json"
    summary_path.write_text(json.dumps(summary_rows, indent=2), encoding="utf-8")
    artifact_paths["summary"] = summary_path

    return artifact_paths


__all__ = [
    "GAS_CONSTANT_J_PER_MOL_K",
    "ILMENITE_ACTIVATION_ENERGIES_J_PER_MOL",
    "ILMENITE_CONVERSION_DECAY",
    "ILMENITE_DIFFUSION_ACTIVATION_ENERGIES_J_PER_MOL",
    "ILMENITE_DIFFUSIVITY_COEFFICIENTS_M2_PER_S",
    "ILMENITE_PARTICLE_RADIUS_M",
    "ILMENITE_RATE_COEFFICIENTS_M_PER_S",
    "ILMENITE_REACTION_STANDARD_HEATS_J_PER_MOL",
    "ILMENITE_REACTION_STOICHIOMETRIES",
    "ILMENITE_SOLID_CONCENTRATIONS_MOL_PER_M3",
    "ILMENITE_WGS_ACTIVATION_ENERGY_J_PER_MOL",
    "ILMENITE_WGS_PREEXPONENTIAL_MOL_PER_G_S",
    "IlmeniteRedoxTerms",
    "catalyst_mass_density_value",
    "co_reduction_site_rate_value",
    "diffusivity_value",
    "equilibrium_constant_wgs_value",
    "gas_phase_concentration_value",
    "h2_reduction_site_rate_value",
    "mixed_control_conversion_rate_value",
    "o2_oxidation_site_rate_value",
    "oxidation_conversion_value",
    "oxygen_capacity_value",
    "partial_pressure_kpa_value",
    "partial_pressure_value",
    "reaction_constant_value",
    "reaction_enthalpy_expression",
    "reaction_enthalpy_value",
    "reduction_conversion_value",
    "run_ortiz_2016_tga_validation",
    "wgs_rate_constant_value",
    "wgs_rate_value",
    "paper_oxidized_bed_profile_mol_per_m3",
    "paper_reduced_bed_profile_mol_per_m3",
    "paper_tga_bed_length_m",
    "paper_tga_bed_volume_m3",
]
